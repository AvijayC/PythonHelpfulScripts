import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from typing import Optional, Dict, List, Tuple, Any
from config import SubtableSearchConfig, ColumnConfig


class SubtableExtractor:
    """Extracts subtables from Excel worksheets based on configurable patterns."""
    
    def __init__(self, workbook: Workbook, debug: bool = False):
        """
        Initialize the extractor with an openpyxl Workbook.
        
        Args:
            workbook: An openpyxl Workbook object
            debug: Enable debug output
        """
        self.workbook = workbook
        self.debug = debug
    
    def extract(self, sheet_name: str, config: SubtableSearchConfig) -> pd.DataFrame:
        """
        Extract one or more subtables from the specified sheet using the provided configuration.

        Args:
            sheet_name: Name of the worksheet to scan
            config: Configuration defining search patterns and extraction rules

        Returns:
            DataFrame with extracted rows and metadata (or list of DataFrames if combine_subtables=False)
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        worksheet = self.workbook[sheet_name]

        # If not extracting multiple, use the single extraction logic
        if not config.extract_multiple:
            return self._extract_single_subtable(worksheet, sheet_name, config, 1, worksheet.max_row)

        # Extract multiple subtables
        all_subtables = []
        current_row = 1
        subtable_index = 0
        consecutive_blank_rows = 0

        if self.debug:
            print(f"Starting multi-subtable extraction from sheet '{sheet_name}'")

        while current_row <= worksheet.max_row:
            # Check if we've reached the maximum number of subtables
            if config.max_subtables and subtable_index >= config.max_subtables:
                if self.debug:
                    print(f"Reached maximum subtable limit ({config.max_subtables})")
                break

            # Look for the next section header
            if config.section_header:
                section_info = self._find_section_header_from_row(
                    worksheet, config.section_header, current_row
                )
                if not section_info:
                    # No more section headers found
                    if self.debug:
                        print(f"No more section headers found after row {current_row}")
                    break

                section_text, section_start_row, section_end_row = section_info
                search_end_row = worksheet.max_row  # Search until end or next header
            else:
                # No section header - try to find next header row
                if subtable_index > 0:
                    # For subsequent subtables without section headers,
                    # look for the next occurrence of column headers
                    header_row = self._find_header_row(worksheet, current_row, config.columns)
                    if not header_row:
                        if self.debug:
                            print(f"No more column headers found after row {current_row}")
                        break
                    # Back up to use the section-finding logic
                    section_text = ""
                    section_start_row = header_row
                    section_end_row = header_row - 1
                    search_end_row = worksheet.max_row
                else:
                    # First subtable without section header
                    section_text = ""
                    section_start_row = 1
                    section_end_row = 0
                    search_end_row = worksheet.max_row

            # Extract this subtable
            subtable_df = self._extract_single_subtable(
                worksheet, sheet_name, config,
                section_end_row + 1, search_end_row,
                section_text=section_text,
                subtable_index=subtable_index
            )

            if not subtable_df.empty:
                all_subtables.append(subtable_df)
                subtable_index += 1
                consecutive_blank_rows = 0

                # Move to the row after this subtable ends
                last_row = subtable_df['row_number'].max()
                current_row = last_row + 1

                if self.debug:
                    print(f"Extracted subtable {subtable_index} with {len(subtable_df)} rows")
            else:
                # No valid subtable found at this position
                consecutive_blank_rows += 1
                current_row = section_end_row + 1 if config.section_header else current_row + 1

                # Check if we've exceeded the maximum blank rows between subtables
                if consecutive_blank_rows >= config.max_blank_rows_between_subtables:
                    if self.debug:
                        print(f"Exceeded max blank rows between subtables ({config.max_blank_rows_between_subtables})")
                    break

        # Combine or return list based on configuration
        if not all_subtables:
            return pd.DataFrame()

        if config.combine_subtables:
            # Combine all subtables into one DataFrame
            combined_df = pd.concat(all_subtables, ignore_index=True)
            if self.debug:
                print(f"Combined {len(all_subtables)} subtables into single DataFrame with {len(combined_df)} total rows")
            return combined_df
        else:
            # Return list of DataFrames (user would need to handle this differently)
            if self.debug:
                print(f"Returning list of {len(all_subtables)} separate DataFrames")
            return all_subtables

    def _extract_single_subtable(self, worksheet, sheet_name: str, config: SubtableSearchConfig,
                                  start_search_row: int, end_search_row: int,
                                  section_text: str = None, subtable_index: int = 0) -> pd.DataFrame:
        """
        Extract a single subtable from the worksheet.

        Args:
            worksheet: The worksheet object
            sheet_name: Name of the worksheet
            config: Configuration for extraction
            start_search_row: Row to start searching from
            end_search_row: Row to stop searching at
            section_text: Text of the section header (if found)
            subtable_index: Index of this subtable (for multi-subtable extraction)

        Returns:
            DataFrame with extracted rows and metadata
        """
        # If section_text wasn't provided, try to find it
        if section_text is None and config.section_header:
            section_info = self._find_section_header_from_row(
                worksheet, config.section_header, start_search_row
            )
            if not section_info:
                if self.debug:
                    print(f"Section header not found starting from row {start_search_row}")
                return pd.DataFrame()
            section_text, section_start_row, section_end_row = section_info
            search_start_row = section_end_row + 1
        else:
            search_start_row = start_search_row
            if section_text is None:
                section_text = ""

        # Find column headers - look for first non-empty row after section
        header_row = self._find_header_row(worksheet, search_start_row, config.columns)
        if not header_row:
            return pd.DataFrame()

        column_mapping = self._find_column_headers(worksheet, header_row, config.columns, config)
        if not column_mapping:
            return pd.DataFrame()

        # Extract rows (limiting to end_search_row)
        rows_data = self._extract_rows(
            worksheet,
            header_row + 1,
            column_mapping,
            config,
            max_row=end_search_row
        )

        # Create DataFrame with metadata
        if not rows_data:
            return pd.DataFrame()

        df = pd.DataFrame(rows_data)

        # Add metadata columns
        df['section_header'] = section_text
        df['sheet_name'] = sheet_name
        df['subtable_index'] = subtable_index  # Add subtable index for tracking

        return df
    
    def _find_section_header(self, worksheet, header_config) -> Optional[Tuple[str, int, int]]:
        """
        Find the section header in the worksheet starting from row 1.

        Returns:
            Tuple of (section_text, start_row, end_row) or None if not found
        """
        return self._find_section_header_from_row(worksheet, header_config, 1)

    def _find_section_header_from_row(self, worksheet, header_config, start_row: int) -> Optional[Tuple[str, int, int]]:
        """
        Find the section header in the worksheet starting from a specific row.

        Args:
            worksheet: The worksheet to search in
            header_config: Configuration for the section header
            start_row: Row number to start searching from

        Returns:
            Tuple of (section_text, start_row, end_row) or None if not found
        """
        col_idx = ord(header_config.start_column.upper()) - ord('A') + 1

        if self.debug:
            print(f"Looking for section header starting at column {header_config.start_column} (idx: {col_idx}), from row {start_row}")
            print(f"Pattern: {header_config.pattern.pattern}, Is merged: {header_config.is_merged}")

        for row in range(start_row, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=col_idx)
            
            if header_config.is_merged:
                # Check if this cell is part of a merged range
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Get the value from the top-left cell of the merged range
                        top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        
                        if self.debug:
                            print(f"  Found merged cell at {merged_range}: '{top_left_cell.value}'")
                        
                        # Check if it starts at the required column
                        if merged_range.min_col != col_idx:
                            continue
                            
                        # Validate merged dimensions if specified
                        if header_config.merged_rows and merged_range.max_row - merged_range.min_row + 1 != header_config.merged_rows:
                            continue
                        if header_config.merged_columns and merged_range.max_col - merged_range.min_col + 1 != header_config.merged_columns:
                            continue
                        
                        if top_left_cell.value and header_config.pattern.match(str(top_left_cell.value)):
                            if self.debug:
                                print(f"  MATCHED section header: '{top_left_cell.value}'")
                            return (str(top_left_cell.value), merged_range.min_row, merged_range.max_row)
            else:
                # Simple cell check
                if cell.value:
                    if self.debug and row <= 5:  # Only debug first few rows
                        print(f"  Row {row}: '{cell.value}'")
                    
                    if header_config.pattern.match(str(cell.value)):
                        if self.debug:
                            print(f"  MATCHED section header: '{cell.value}'")
                        return (str(cell.value), row, row)
        
        if self.debug:
            print("  No section header found")
        return None
    
    def _find_header_row(self, worksheet, start_row: int, columns: List[ColumnConfig]) -> Optional[int]:
        """
        Find the row containing column headers by scanning for a row that matches at least one header pattern.
        
        Returns:
            Row number of the header row, or None if not found
        """
        max_search_rows = 10  # Don't search more than 10 rows for headers
        
        for row_offset in range(max_search_rows):
            row = start_row + row_offset
            if row > worksheet.max_row:
                break
            
            # Check if this row contains any matching headers
            matches_found = 0
            for col_config in columns:
                col_idx = ord(col_config.column_letter.upper()) - ord('A') + 1
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value and col_config.header_pattern.match(str(cell.value).strip()):
                    matches_found += 1
            
            # If we found at least one matching header, this is likely the header row
            if matches_found > 0:
                if self.debug:
                    print(f"Found header row at row {row} with {matches_found} matching headers")
                return row
        
        return None
    
    def _find_column_headers(self, worksheet, header_row: int, columns: List[ColumnConfig], config: SubtableSearchConfig) -> Dict[str, Dict[str, Any]]:
        """
        Find and map column headers based on configuration.

        Returns:
            Dictionary mapping column names to their info (index, config)
        """
        column_mapping = {}

        if self.debug:
            print(f"\nLooking for column headers at row {header_row}")

        # First, handle fixed-position columns (existing logic)
        for col_config in columns:
            col_idx = ord(col_config.column_letter.upper()) - ord('A') + 1
            cell = worksheet.cell(row=header_row, column=col_idx)

            if self.debug:
                print(f"  Column {col_config.column_letter}: '{cell.value}' (pattern: {col_config.header_pattern.pattern})")

            if cell.value and col_config.header_pattern.match(str(cell.value).strip()):
                column_mapping[col_config.column_letter] = {
                    'index': col_idx,
                    'header_text': str(cell.value).strip(),
                    'config': col_config
                }
                if self.debug:
                    print(f"    MATCHED!")

        # ADD THIS SECTION - Discover additional columns dynamically
        if config.discoverable_headers:
            if self.debug:
                print(f"\n  Searching for discoverable headers...")

            # Scan entire row for discoverable patterns
            for col_idx in range(1, worksheet.max_column + 1):
                col_letter = chr(ord('A') + col_idx - 1) if col_idx <= 26 else f"{chr(ord('A') + (col_idx - 1) // 26 - 1)}{chr(ord('A') + (col_idx - 1) % 26)}"

                # Skip if already mapped as fixed column
                if col_letter in column_mapping:
                    continue

                cell = worksheet.cell(row=header_row, column=col_idx)
                if cell.value:
                    cell_text = str(cell.value).strip()

                    # Check against each discoverable pattern
                    for pattern in config.discoverable_headers:
                        if pattern.match(cell_text):
                            if self.debug:
                                print(f"    Discovered column {col_letter}: '{cell_text}' matches pattern {pattern.pattern}")

                            # Create a dynamic ColumnConfig for this discovered column
                            dynamic_config = ColumnConfig(
                                column_letter=col_letter,
                                header_pattern=pattern,
                                value_pattern=re.compile(r".*")  # Accept any value by default
                            )

                            column_mapping[col_letter] = {
                                'index': col_idx,
                                'header_text': cell_text,
                                'config': dynamic_config,
                                'is_discovered': True  # Mark as dynamically discovered
                            }
                            break  # Stop checking patterns once matched

        if self.debug:
            print(f"  Found {len(column_mapping)} column headers total")

        # Strict column validation - check for unexpected columns
        if config.strict_columns:
            unexpected_columns = []

            for col_idx in range(1, worksheet.max_column + 1):
                # Convert column index to letter (handles A-Z and AA-AZ)
                col_letter = chr(ord('A') + col_idx - 1) if col_idx <= 26 else f"{chr(ord('A') + (col_idx - 1) // 26 - 1)}{chr(ord('A') + (col_idx - 1) % 26)}"

                # Skip if this column is already mapped (either fixed or discovered)
                if col_letter in column_mapping:
                    continue

                cell = worksheet.cell(row=header_row, column=col_idx)
                if cell.value and str(cell.value).strip():
                    # Found a non-empty, unmapped column
                    unexpected_columns.append(f"{col_letter}: '{str(cell.value).strip()}'")

            if unexpected_columns:
                raise ValueError(
                    f"Unexpected columns found in header row {header_row}: {', '.join(unexpected_columns)}. "
                    f"Either add them to 'columns' config, 'discoverable_headers' patterns, or set strict_columns=False"
                )

        return column_mapping
    
    def _extract_rows(self, worksheet, start_row: int, column_mapping: Dict, config: SubtableSearchConfig, max_row: int = None) -> List[Dict]:
        """
        Extract valid rows from the worksheet based on configuration.

        Args:
            worksheet: The worksheet to extract from
            start_row: Row to start extraction from
            column_mapping: Column mapping dictionary
            config: Extraction configuration
            max_row: Maximum row to extract to (None = worksheet.max_row)

        Returns:
            List of dictionaries representing valid rows with metadata
        """
        rows_data = []
        current_row = start_row
        consecutive_invalid_rows = 0
        consecutive_blank_rows = 0

        if max_row is None:
            max_row = worksheet.max_row

        while current_row <= max_row:
            if self.debug:
                print(f"  Processing row {current_row}")
            
            # Check end conditions
            end_check = self._check_end_condition(worksheet, current_row, column_mapping, config)
            if end_check:
                if self.debug:
                    print(f"    End condition met at row {current_row}")
                break
            
            # Extract row data
            row_data = self._extract_single_row(worksheet, current_row, column_mapping)
            
            # Check if it's a blank row
            is_blank = all(not row_data.get(col_info['header_text']) for col_info in column_mapping.values())
            
            # Validate row
            if self._is_valid_row(row_data, column_mapping, config):
                # Valid row - add metadata
                row_data['row_start'] = f"{list(column_mapping.keys())[0]}{current_row}"
                row_data['row_end'] = f"{list(column_mapping.keys())[-1]}{current_row}"
                row_data['row_number'] = current_row
                
                rows_data.append(row_data)
                
                # Reset all counters for valid rows
                consecutive_invalid_rows = 0
                consecutive_blank_rows = 0
                
                if self.debug:
                    print(f"    Valid row added")
            else:
                # Invalid row (includes blank rows)
                consecutive_invalid_rows += 1
                
                if is_blank:
                    consecutive_blank_rows += 1
                    if self.debug:
                        print(f"    Blank row (consecutive invalid: {consecutive_invalid_rows}, consecutive blank: {consecutive_blank_rows})")
                else:
                    consecutive_blank_rows = 0  # Reset blank counter for non-blank invalid rows
                    if self.debug:
                        print(f"    Invalid row (consecutive invalid: {consecutive_invalid_rows})")
                
                # Check stop conditions
                if config.max_consecutive_invalid_rows > 0 and consecutive_invalid_rows >= config.max_consecutive_invalid_rows:
                    if self.debug:
                        print(f"    Stopping: reached {config.max_consecutive_invalid_rows} consecutive invalid rows")
                    break
                if config.max_consecutive_blank_rows > 0 and consecutive_blank_rows >= config.max_consecutive_blank_rows:
                    if self.debug:
                        print(f"    Stopping: reached {config.max_consecutive_blank_rows} consecutive blank rows")
                    break
            
            current_row += 1
        
        return rows_data
    
    def _extract_cell_value_with_type(self, cell) -> Any:
        """
        Extract cell value preserving the correct Python type based on Excel metadata.

        This method intelligently converts Excel cell values to appropriate Python types
        based on the cell's data_type, number_format, and other metadata.

        Returns:
            The cell value with the appropriate Python type (int, float, str, datetime, bool, None)
        """
        if cell.value is None:
            return None

        # String/Text cells (including text-formatted numbers like '100)
        if cell.data_type == 's' or cell.data_type in ['str', 'inlineStr']:
            return str(cell.value)

        # Date cells
        elif cell.data_type == 'd' or cell.is_date:
            return cell.value  # Already a datetime object

        # Numeric cells
        elif cell.data_type == 'n':
            val = cell.value

            if val is None:
                return 0.0

            # Safe integer conversion for all whole numbers
            if isinstance(val, (int, float)):
                # Check if it's a whole number
                if val == int(val):
                    # Check if it's within safe integer range
                    # Using 2^53 as the safe boundary (same as JavaScript's Number.MAX_SAFE_INTEGER)
                    # This ensures compatibility across systems and prevents precision issues
                    if -2**53 <= val <= 2**53:
                        return int(val)
                    # For very large whole numbers, keep as float to maintain precision

            # Return as float for non-whole numbers or very large numbers
            return float(val) if not isinstance(val, float) else val

        # Boolean cells
        elif cell.data_type == 'b':
            return bool(cell.value)

        # Formula cells (with data_only=True, we get the calculated value)
        elif cell.data_type == 'f':
            # The value type depends on the formula result
            # Recurse with a synthetic cell-like object
            class CellLike:
                def __init__(self, value, data_type='n', number_format='General', is_date=False):
                    self.value = value
                    self.data_type = data_type
                    self.number_format = number_format
                    self.is_date = is_date

            # Try to infer the type from the value
            if isinstance(cell.value, bool):
                return cell.value
            elif isinstance(cell.value, (int, float)):
                synthetic = CellLike(cell.value, 'n', cell.number_format, cell.is_date)
                return self._extract_cell_value_with_type(synthetic)
            else:
                return cell.value

        # Error cells
        elif cell.data_type == 'e':
            return f"#ERROR: {cell.value}"

        # Fallback - preserve as-is
        else:
            return cell.value

    def _extract_single_row(self, worksheet, row_num: int, column_mapping: Dict) -> Dict:
        """Extract data from a single row with proper type preservation."""
        row_data = {}

        for col_letter, col_info in column_mapping.items():
            cell = worksheet.cell(row=row_num, column=col_info['index'])

            # Use the new type-preserving extraction method
            value = self._extract_cell_value_with_type(cell)

            # Store the properly typed value
            row_data[col_info['header_text']] = value if value is not None else ""

            # Store cell coordinate for reference
            row_data[f"{col_info['header_text']}_coord"] = f"{col_letter}{row_num}"

        return row_data
    
    def _is_valid_row(self, row_data: Dict, column_mapping: Dict, config: SubtableSearchConfig) -> bool:
        """
        Determine if a row is valid based on configuration.
        
        Returns:
            True if the row meets validation criteria
        """
        non_empty_columns = 0
        
        # Check each column's value against its regex pattern
        for col_letter, col_info in column_mapping.items():
            col_config = col_info['config']
            header_text = col_info['header_text']
            value = row_data.get(header_text, "")
            
            # Check if value matches the pattern (including empty strings)
            if not col_config.value_pattern.match(str(value)):
                if self.debug:
                    print(f"    Row invalid: value '{value}' doesn't match pattern {col_config.value_pattern.pattern}")
                return False
            
            # Count non-empty columns
            if value:
                non_empty_columns += 1
        
        # Check minimum filled columns if specified
        if config.row_validation.minimum_filled_columns > 0:
            if non_empty_columns < config.row_validation.minimum_filled_columns:
                if self.debug:
                    print(f"    Row invalid: only {non_empty_columns} filled, need {config.row_validation.minimum_filled_columns}")
                return False
        
        return True
    
    def _check_end_condition(self, worksheet, row_num: int, column_mapping: Dict, config: SubtableSearchConfig) -> bool:
        """
        Check if the end condition for the subtable has been met.
        
        Returns:
            True if end condition is met
        """
        # Check for merged cells if configured
        if config.stop_on_merged_cell:
            for col_info in column_mapping.values():
                cell = worksheet.cell(row=row_num, column=col_info['index'])
                if isinstance(cell, MergedCell) or any(cell.coordinate in r for r in worksheet.merged_cells.ranges):
                    if self.debug:
                        print(f"    Merged cell detected at row {row_num}")
                    return True
        
        # Check custom end pattern if provided
        if config.end_pattern and config.end_pattern_column:
            col_idx = ord(config.end_pattern_column.upper()) - ord('A') + 1
            cell = worksheet.cell(row=row_num, column=col_idx)
            if cell.value and config.end_pattern.match(str(cell.value)):
                if self.debug:
                    print(f"    End pattern matched at row {row_num}: '{cell.value}'")
                return True
        
        return False