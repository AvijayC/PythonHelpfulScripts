#!/usr/bin/env python3
"""Create a simple test Excel file"""

from openpyxl import Workbook

def create_test_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    
    # Add some test data
    ws['A1'] = "Sample Data Report"  # Section header
    ws['A2'] = ""  # Empty row
    ws['A3'] = "colA"  # Column headers
    ws['B3'] = "colB"
    ws['C3'] = "colC"
    
    # Add some test data
    ws['A4'] = "value1"
    ws['B4'] = "value2" 
    ws['C4'] = "value3"
    
    ws['A5'] = "test1"
    ws['B5'] = ""  # Empty cell
    ws['C5'] = "test3"
    
    ws['A6'] = ""  # Partially empty row
    ws['B6'] = "only_b"
    ws['C6'] = ""
    
    # Two blank rows to trigger stop condition
    # Rows 7-8 are empty
    
    wb.save("simple_test.xlsx")
    print("Created simple_test.xlsx with:")
    print("  Row 1: 'Sample Data Report' (section header)")
    print("  Row 3: colA, colB, colC (headers)")
    print("  Rows 4-6: test data")
    print("  Rows 7-8: blank (should trigger stop)")

if __name__ == "__main__":
    create_test_file()