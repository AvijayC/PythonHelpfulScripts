"""
Comprehensive test to validate extractor output against actual Excel data.
This script will extract data and compare it with what's actually in the Excel file.
"""

import re
import json
from openpyxl import load_workbook
from extractor import SubtableExtractor
from config import (
    SubtableSearchConfig, 
    SectionHeaderConfig, 
    ColumnConfig, 
    RowValidationConfig,
    EndCondition
)


def get_actual_excel_data(wb, sheet_name, start_row, end_row, columns):
    """Get actual data from Excel for comparison."""
    ws = wb[sheet_name]
    data = []
    
    for row in range(start_row, end_row + 1):
        row_data = {}
        for col_letter in columns:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=row, column=col_idx)
            row_data[col_letter] = str(cell.value) if cell.value else ""
        data.append(row_data)
    
    return data


def test_normal_cases():
    """Test extraction from Normal Cases sheet."""
    print("=" * 70)
    print("TEST: Normal Cases Sheet - Employee Records")
    print("=" * 70)
    
    wb = load_workbook("comprehensive_test_workbook.xlsx", data_only=True)
    ws = wb["Normal Cases"]
    
    # Show actual Excel data
    print("\n### ACTUAL EXCEL DATA (Rows 1-11):")
    print("-" * 50)
    for row in range(1, 12):
        row_data = []
        for col in range(1, 6):  # Columns A-E
            cell = ws.cell(row=row, column=col)
            value = str(cell.value)[:20] if cell.value else "<empty>"
            row_data.append(f"{chr(64+col)}: {value}")
        print(f"Row {row:2}: {' | '.join(row_data[:3])}")  # Show first 3 columns for brevity
    
    # Configure and extract
    extractor = SubtableExtractor(wb, debug=False)
    config = SubtableSearchConfig(
        section_header=SectionHeaderConfig(
            pattern=re.compile(r".*Employee.*", re.IGNORECASE),
            start_column="A",
            is_merged=True
        ),
        columns=[
            ColumnConfig(
                column_letter="A",
                header_pattern=re.compile(r".*ID.*", re.IGNORECASE),
                value_pattern=re.compile(r"EMP-\d{3}")  # Required: exact format
            ),
            ColumnConfig(
                column_letter="B",
                header_pattern=re.compile(r".*Name.*", re.IGNORECASE),
                value_pattern=re.compile(r".+")  # Required: non-empty
            ),
            ColumnConfig(
                column_letter="C",
                header_pattern=re.compile(r".*Email.*", re.IGNORECASE),
                value_pattern=re.compile(r".*@.*")  # Required: contains @
            ),
            ColumnConfig(
                column_letter="E",
                header_pattern=re.compile(r".*Salary.*", re.IGNORECASE),
                value_pattern=re.compile(r"\$[\d,]+\.?\d*")  # Required: money format
            ),
            ColumnConfig(
                column_letter="D",
                header_pattern=re.compile(r".*Phone.*", re.IGNORECASE),
                value_pattern=re.compile(r".+")  # Required: money format
            )
        ],
        row_validation=RowValidationConfig(
            minimum_filled_columns=3
        ),
        end_condition=EndCondition.FIRST_BLANK_ROW
    )
    
    df = extractor.extract("Normal Cases", config)
    
    print("\n### EXTRACTED DATAFRAME:")
    print("-" * 50)
    if not df.empty:
        # Show extracted data
        data_cols = ['Employee ID', 'Full Name', 'Email Address', 'Salary']
        for idx, row in df.iterrows():
            print(f"Row {idx+1}: ID={row['Employee ID']}, Name={row['Full Name']}, "
                  f"Email={row['Email Address'][:20]}..., Salary={row['Salary']}")
        
        print(f"\nTotal rows extracted: {len(df)}")
        print(f"Row numbers in Excel: {df['row_number'].tolist()}")
    else:
        print("No data extracted!")
    
    return {
        "sheet": "Normal Cases",
        "section": "Employee Records",
        "expected_rows": [6, 7, 8, 9],
        "actual_rows": df['row_number'].tolist() if not df.empty else [],
        "dataframe": df
    }


def test_edge_cases():
    """Test extraction from Edge Cases sheet."""
    print("\n" + "=" * 70)
    print("TEST: Edge Cases Sheet - Product Inventory")
    print("=" * 70)
    
    wb = load_workbook("comprehensive_test_workbook.xlsx", data_only=True)
    ws = wb["Edge Cases"]
    
    # Show actual Excel data
    print("\n### ACTUAL EXCEL DATA (Rows 1-10):")
    print("-" * 50)
    for row in range(1, 11):
        row_data = []
        for col in range(1, 4):  # Columns A-C
            cell = ws.cell(row=row, column=col)
            value = str(cell.value)[:15] if cell.value else "<empty>"
            row_data.append(f"{value}")
        print(f"Row {row:2}: {' | '.join(row_data)}")
    
    # Configure and extract
    extractor = SubtableExtractor(wb, debug=False)
    config = SubtableSearchConfig(
        section_header=SectionHeaderConfig(
            pattern=re.compile(r".*Product.*Inventory.*", re.IGNORECASE),
            start_column="A",
            is_merged=False
        ),
        columns=[
            ColumnConfig(
                column_letter="A",
                header_pattern=re.compile(r".*Product.*ID.*", re.IGNORECASE),
                value_pattern=re.compile(r"PRD-\d{3}")  # Required format
            ),
            ColumnConfig(
                column_letter="B",
                header_pattern=re.compile(r".*NAME.*", re.IGNORECASE),
                value_pattern=re.compile(r".*")  # Optional
            ),
            ColumnConfig(
                column_letter="C",
                header_pattern=re.compile(r".*SKU.*", re.IGNORECASE),
                value_pattern=re.compile(r"^$|SKU\d{5}")  # Optional: empty or SKU format
            )
        ],
        row_validation=RowValidationConfig(
            minimum_filled_columns=1
        ),
        end_condition=EndCondition.FIRST_BLANK_ROW,
        max_blank_rows=1  # Allow one blank row
    )
    
    df = extractor.extract("Edge Cases", config)
    
    print("\n### EXTRACTED DATAFRAME:")
    print("-" * 50)
    if not df.empty:
        for idx, row in df.iterrows():
            print(f"Row {idx+1}: ID={row['Product ID']}, "
                  f"Name={'<empty>' if not row['product NAME'] else row['product NAME'][:10]}, "
                  f"SKU={'<empty>' if not row['SKU_Code'] else row['SKU_Code']}")
        
        print(f"\nTotal rows extracted: {len(df)}")
        print(f"Row numbers in Excel: {df['row_number'].tolist()}")
    else:
        print("No data extracted!")
    
    return {
        "sheet": "Edge Cases",
        "section": "Product Inventory",
        "expected_rows": [4, 5, 6, 7, 8, 10],  # Row 9 is blank
        "actual_rows": df['row_number'].tolist() if not df.empty else [],
        "dataframe": df
    }


def test_validation_tests():
    """Test extraction from Validation Tests sheet."""
    print("\n" + "=" * 70)
    print("TEST: Validation Tests Sheet - Transaction Log")
    print("=" * 70)
    
    wb = load_workbook("comprehensive_test_workbook.xlsx", data_only=True)
    ws = wb["Validation Tests"]
    
    # Show actual Excel data
    print("\n### ACTUAL EXCEL DATA (Rows 1-11):")
    print("-" * 50)
    for row in range(1, 12):
        row_data = []
        for col in range(1, 5):  # Columns A-D
            cell = ws.cell(row=row, column=col)
            value = str(cell.value)[:15] if cell.value else "<empty>"
            row_data.append(f"{value}")
        print(f"Row {row:2}: {' | '.join(row_data)}")
    
    # Configure and extract - strict validation
    extractor = SubtableExtractor(wb, debug=False)
    config = SubtableSearchConfig(
        section_header=SectionHeaderConfig(
            pattern=re.compile(r".*Transaction.*Log.*", re.IGNORECASE),
            start_column="A",
            is_merged=True
        ),
        columns=[
            ColumnConfig(
                column_letter="A",
                header_pattern=re.compile(r".*Transaction.*ID.*", re.IGNORECASE),
                value_pattern=re.compile(r"TXN-\d{3}")  # Strict format
            ),
            ColumnConfig(
                column_letter="B",
                header_pattern=re.compile(r".*Date.*", re.IGNORECASE),
                value_pattern=re.compile(r"\d{4}-\d{2}-\d{2}")  # YYYY-MM-DD format
            ),
            ColumnConfig(
                column_letter="C",
                header_pattern=re.compile(r".*Amount.*", re.IGNORECASE),
                value_pattern=re.compile(r"\$[\d,]+\.?\d*")  # Money format
            ),
            ColumnConfig(
                column_letter="D",
                header_pattern=re.compile(r".*Status.*", re.IGNORECASE),
                value_pattern=re.compile(r"APPROVED|PENDING|REJECTED")  # Exact statuses
            )
        ],
        row_validation=RowValidationConfig(
            minimum_filled_columns=4  # All columns required
        ),
        end_condition=EndCondition.FIRST_INVALID_ROW
    )
    
    df = extractor.extract("Validation Tests", config)
    
    print("\n### EXTRACTED DATAFRAME (Strict Validation):")
    print("-" * 50)
    if not df.empty:
        for idx, row in df.iterrows():
            print(f"Row {idx+1}: ID={row['Transaction ID']}, Date={row['Date']}, "
                  f"Amount={row['Amount']}, Status={row['Status']}")
        
        print(f"\nTotal rows extracted: {len(df)}")
        print(f"Row numbers in Excel: {df['row_number'].tolist()}")
    else:
        print("No data extracted with strict validation!")
    
    return {
        "sheet": "Validation Tests",
        "section": "Transaction Log",
        "expected_valid_rows": [5, 6, 8],  # Only rows with correct format
        "actual_rows": df['row_number'].tolist() if not df.empty else [],
        "dataframe": df
    }


def test_multiple_tables():
    """Test extraction from Multiple Tables sheet."""
    print("\n" + "=" * 70)
    print("TEST: Multiple Tables Sheet - Sales Data")
    print("=" * 70)
    
    wb = load_workbook("comprehensive_test_workbook.xlsx", data_only=True)
    ws = wb["Multiple Tables"]
    
    # Show actual Excel data
    print("\n### ACTUAL EXCEL DATA (All rows):")
    print("-" * 50)
    for row in range(1, min(15, ws.max_row + 1)):
        row_data = []
        for col in range(1, 4):  # Columns A-C
            cell = ws.cell(row=row, column=col)
            value = str(cell.value)[:15] if cell.value else "<empty>"
            row_data.append(f"{value}")
        print(f"Row {row:2}: {' | '.join(row_data)}")
    
    # Configure and extract first table
    extractor = SubtableExtractor(wb, debug=False)
    config = SubtableSearchConfig(
        section_header=SectionHeaderConfig(
            pattern=re.compile(r"Sales Data.*Region North", re.IGNORECASE),
            start_column="A",
            is_merged=True
        ),
        columns=[
            ColumnConfig(
                column_letter="A",
                header_pattern=re.compile(r".*Sale.*ID.*", re.IGNORECASE),
                value_pattern=re.compile(r"SALE-\d{3}")
            ),
            ColumnConfig(
                column_letter="B",
                header_pattern=re.compile(r".*Product.*", re.IGNORECASE),
                value_pattern=re.compile(r"Product [A-Z]")
            ),
            ColumnConfig(
                column_letter="C",
                header_pattern=re.compile(r".*Amount.*", re.IGNORECASE),
                value_pattern=re.compile(r"\d+")
            )
        ],
        row_validation=RowValidationConfig(
            minimum_filled_columns=3
        ),
        end_condition=EndCondition.FIRST_BLANK_ROW
    )
    
    df = extractor.extract("Multiple Tables", config)
    
    print("\n### EXTRACTED DATAFRAME (First Table - North Region):")
    print("-" * 50)
    if not df.empty:
        for idx, row in df.iterrows():
            print(f"Row {idx+1}: ID={row['Sale ID']}, Product={row['Product']}, Amount={row['Amount']}")
        
        print(f"\nTotal rows extracted: {len(df)}")
        print(f"Row numbers in Excel: {df['row_number'].tolist()}")
    else:
        print("No data extracted!")
    
    return {
        "sheet": "Multiple Tables",
        "section": "Sales Data - Region North",
        "expected_rows": [4, 5],  # Two data rows before blank
        "actual_rows": df['row_number'].tolist() if not df.empty else [],
        "dataframe": df
    }


def main():
    """Run all tests and generate report."""
    results = []
    
    # Run all tests
    results.append(test_normal_cases())
    results.append(test_edge_cases())
    results.append(test_validation_tests())
    results.append(test_multiple_tables())
    
    # Generate summary
    print("\n" + "=" * 70)
    print("SUMMARY OF ALL TESTS")
    print("=" * 70)
    
    for result in results:
        print(f"\n{result['sheet']} - {result['section']}:")
        print(f"  Expected rows: {result.get('expected_rows', result.get('expected_valid_rows', []))}")
        print(f"  Actual rows:   {result['actual_rows']}")
        match = result['actual_rows'] == result.get('expected_rows', result.get('expected_valid_rows', []))
        print(f"  Match: {'✅ YES' if match else '❌ NO'}")
    
    return results


if __name__ == "__main__":
    results = main()