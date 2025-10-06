#!/usr/bin/env python3
"""Simple test for basic config"""

import re
import pandas as pd
from openpyxl import load_workbook
from extractor import SubtableExtractor
from config import (
    SubtableSearchConfig,
    SectionHeaderConfig, 
    ColumnConfig,
    RowValidationConfig
)

def main():
    # Simple test config
    config = SubtableSearchConfig(
        # Look for any section header containing "data" or "table" (case insensitive)
        section_header=SectionHeaderConfig(
            pattern=re.compile(r".*(data|table).*", re.IGNORECASE),
            start_column="A",
            is_merged=False  # Simple single cell header
        ),
        
        # Three columns with exact header names
        columns=[
            ColumnConfig(
                column_letter="A",
                header_pattern=re.compile(r"colA", re.IGNORECASE),
                value_pattern=re.compile(r".*")  # Accept anything (including empty)
            ),
            ColumnConfig(
                column_letter="B", 
                header_pattern=re.compile(r"colB", re.IGNORECASE),
                value_pattern=re.compile(r".*")  # Accept anything
            ),
            ColumnConfig(
                column_letter="C",
                header_pattern=re.compile(r"colC", re.IGNORECASE), 
                value_pattern=re.compile(r".*")  # Accept anything
            )
        ],
        
        # Minimal validation
        row_validation=RowValidationConfig(
            minimum_filled_columns=1  # At least 1 column must have data
        ),
        
        # Stop conditions - allow some gaps
        max_consecutive_invalid_rows=5,  # Allow up to 5 bad rows in a row
        max_consecutive_blank_rows=2     # Stop after 2 blank rows
    )
    
    print("Simple Test Config:")
    print(f"  Looking for section with: {config.section_header.pattern.pattern}")
    print(f"  Looking for headers: colA, colB, colC")
    print(f"  Accept any values (including empty)")
    print("="*50)
    
    # Test with your workbook - you'll need to provide the path
    WORKBOOK_PATH = input("Enter Excel file path: ").strip()
    if not WORKBOOK_PATH:
        print("No file provided, exiting")
        return
        
    SHEET_NAME = input("Enter sheet name (or press Enter for first sheet): ").strip()
    
    try:
        wb = load_workbook(WORKBOOK_PATH, data_only=True)
        
        if not SHEET_NAME:
            SHEET_NAME = wb.sheetnames[0]  # Use first sheet
            
        print(f"\nLoaded workbook: {WORKBOOK_PATH}")
        print(f"Available sheets: {wb.sheetnames}")
        print(f"Using sheet: {SHEET_NAME}")
        print("="*50)
        
        # Create extractor with debug mode
        extractor = SubtableExtractor(wb, debug=True)
        
        # Run extraction
        df_extracted = extractor.extract(SHEET_NAME, config)
        
        print("\n" + "="*50)
        print(f"RESULT: Found {len(df_extracted)} rows")
        
        if not df_extracted.empty:
            print(f"Extracted from Excel rows: {df_extracted['row_number'].tolist()}")
            print(f"Section: {df_extracted['section_header'].iloc[0]}")
            
            # Show just the data columns
            data_cols = ['colA', 'colB', 'colC']  # Expected column names
            actual_data_cols = [col for col in df_extracted.columns if col not in 
                              ['row_start', 'row_end', 'row_number', 'section_header', 'sheet_name']]
            
            print(f"\nActual column names found: {actual_data_cols}")
            print("\nExtracted Data:")
            print(df_extracted[actual_data_cols].to_string(index=False))
        else:
            print("No data was extracted!")
            print("\nPossible issues:")
            print("- Section header pattern didn't match")
            print("- Column headers (colA, colB, colC) not found")
            print("- No valid data rows")
            
    except Exception as e:
        print(f"Error: {e}")
        print("Make sure the file path is correct and the file is not open in Excel")

if __name__ == "__main__":
    main()