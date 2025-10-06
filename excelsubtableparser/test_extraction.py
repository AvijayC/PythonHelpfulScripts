#!/usr/bin/env python3
"""Test the extractor with the main test case from the notebook."""

import re
import pandas as pd
from openpyxl import load_workbook
from extractor import SubtableExtractor
from config import (
    SubtableSearchConfig, 
    SectionHeaderConfig, 
    ColumnConfig, 
    RowValidationConfig
)

def main():
    # Load the workbook
    WORKBOOK_PATH = "comprehensive_test_workbook copy.xlsx"
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    
    print(f"Workbook loaded: {WORKBOOK_PATH}")
    print(f"Available sheets: {wb.sheetnames}")
    print("\n" + "="*70)
    
    # Configuration
    SHEET_NAME = "Normal Cases 2"
    
    config = SubtableSearchConfig(
        # Section header configuration
        section_header=SectionHeaderConfig(
            pattern=re.compile(r".*Employee.*", re.IGNORECASE),
            start_column="A",
            is_merged=True,
            merged_rows=3,
            merged_columns=6
        ),
        
        # Column configurations
        columns=[
            ColumnConfig(
                column_letter="A",
                header_pattern=re.compile(r".*ID.*", re.IGNORECASE),
                value_pattern=re.compile(r"EMP-\d{3}")  # Required: exact format
            ),
            ColumnConfig(
                column_letter="B",
                header_pattern=re.compile(r".*Name.*", re.IGNORECASE),
                value_pattern=re.compile(r".+")  # Required: at least one character
            ),
            # Skip column C (Email) - example of non-contiguous extraction
            ColumnConfig(
                column_letter="D",
                header_pattern=re.compile(r".*Phone.*", re.IGNORECASE),
                value_pattern=re.compile(r"[\d\-]+")  # Required: phone format (will skip rows with empty phone)
            ),
            ColumnConfig(
                column_letter="E",
                header_pattern=re.compile(r".*Salary.*", re.IGNORECASE),
                value_pattern=re.compile(r"\$[\d,]+\.?\d*")  # Required: money format
            )
        ],
        
        # Row validation
        row_validation=RowValidationConfig(
            minimum_filled_columns=3  # At least 3 columns must have data
        ),
        
        # Stop conditions
        max_consecutive_invalid_rows=3,  # Never stop on invalid rows (0 = never)
        max_consecutive_blank_rows=2, # Stop after 1 blank row
        end_pattern=re.compile(r'afsada', re.IGNORECASE),
        end_pattern_column='A'
    )
    
    print("Configuration defined:")
    print(f"  Target sheet: {SHEET_NAME}")
    print(f"  Section pattern: {config.section_header.pattern.pattern}")
    print(f"  Columns to extract: {[col.column_letter for col in config.columns]}")
    print(f"  Max consecutive invalid rows: {config.max_consecutive_invalid_rows} (0 = never)")
    print(f"  Max consecutive blank rows: {config.max_consecutive_blank_rows} (0 = never)")
    print("\n" + "="*70)
    
    # Create extractor with debug mode
    extractor = SubtableExtractor(wb, debug=True)
    
    # Run extraction
    print(f"Extracting from sheet: {SHEET_NAME}")
    print("="*70)
    
    df_extracted = extractor.extract(SHEET_NAME, config)
    
    print("\n" + "="*70)
    print(f"Extraction complete! Found {len(df_extracted)} rows")
    
    if not df_extracted.empty:
        print(f"\nExtracted from Excel rows: {df_extracted['row_number'].tolist()}")
        print(f"Section: {df_extracted['section_header'].iloc[0]}")
        
        # Get data columns (exclude metadata)
        metadata_cols = ['row_start', 'row_end', 'row_number', 'section_header', 'sheet_name']
        data_cols = [col for col in df_extracted.columns if col not in metadata_cols]
        
        print("\n" + "="*70)
        print("EXTRACTED DATA:")
        print("="*70)
        
        # Display data in ASCII table format
        print(df_extracted[data_cols].to_string(index=False))
        
        print("\n" + "="*70)
        print("FULL DATAFRAME WITH METADATA:")
        print("="*70)
        
        # Show full dataframe with all columns
        print(df_extracted.to_string(index=False))
    else:
        print("\nNo data was extracted!")

if __name__ == "__main__":
    main()